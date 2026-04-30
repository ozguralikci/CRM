from __future__ import annotations

import csv
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import joinedload

from crm_app.database.session import get_session
from crm_app.models import Company, Contact, FieldDefinition
from crm_app.services.company_service import create_company, update_company
from crm_app.services.contact_service import create_contact, update_contact
from crm_app.services.field_service import get_field_values, list_field_definitions


@dataclass
class ImportSummary:
    added: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


COMPANY_COLUMN_MAP = {
    "sirket": "name",
    "sirketadi": "name",
    "sirketadiunvani": "name",
    "ad": "name",
    "name": "name",
    "ulke": "country",
    "country": "country",
    "sehir": "city",
    "city": "city",
    "website": "website",
    "websitesi": "website",
    "web": "website",
    "linkedin": "linkedin",
    "oncelik": "priority",
    "priority": "priority",
}

CONTACT_COLUMN_MAP = {
    "adsoyad": "name",
    "kisi": "name",
    "kisiadi": "name",
    "ad": "name",
    "name": "name",
    "sirket": "company_name",
    "sirketadi": "company_name",
    "company": "company_name",
    "companyname": "company_name",
    "companyid": "company_id",
    "sirketid": "company_id",
    "unvan": "title",
    "title": "title",
    "email": "email",
    "eposta": "email",
    "telefon": "phone",
    "phone": "phone",
    "linkedin": "linkedin",
}


def import_companies(file_path: str) -> ImportSummary:
    summary = ImportSummary()
    try:
        source_rows = _read_tabular_file(file_path)
    except Exception as exc:
        summary.errors.append(f"Dosya okunamadi: {exc}")
        return summary

    definitions = list_field_definitions("company")
    definition_lookup = _build_definition_lookup(definitions)

    with get_session() as session:
        existing_companies = list(session.scalars(select(Company)))
        company_map = {company.name.strip().lower(): company for company in existing_companies if company.name}

    for row_number, row in enumerate(source_rows, start=2):
        try:
            mapped, custom_values = _map_row(row, COMPANY_COLUMN_MAP, definition_lookup)
            name = _clean_text(mapped.get("name"))
            if not name:
                summary.skipped += 1
                summary.errors.append(f"Satir {row_number}: Sirket adi zorunludur.")
                continue

            payload = _build_company_payload(mapped, custom_values)
            existing = company_map.get(name.lower())

            if existing:
                merged_custom_values = get_field_values("company", existing.id)
                merged_custom_values.update(payload.pop("custom_values"))
                payload["custom_values"] = merged_custom_values
                update_company(existing.id, payload)
                summary.updated += 1
            else:
                company = create_company(payload)
                company_map[name.lower()] = company
                summary.added += 1
        except Exception as exc:
            summary.errors.append(f"Satir {row_number}: {exc}")

    return summary


def import_contacts(file_path: str) -> ImportSummary:
    summary = ImportSummary()
    try:
        source_rows = _read_tabular_file(file_path)
    except Exception as exc:
        summary.errors.append(f"Dosya okunamadi: {exc}")
        return summary

    definitions = list_field_definitions("contact")
    definition_lookup = _build_definition_lookup(definitions)

    with get_session() as session:
        companies = list(session.scalars(select(Company)))
        company_by_name = {company.name.strip().lower(): company for company in companies if company.name}
        company_by_id = {company.id: company for company in companies}
        contacts = list(
            session.scalars(select(Contact).options(joinedload(Contact.company))).unique()
        )
        contact_map = {
            (contact.company_id, contact.name.strip().lower()): contact
            for contact in contacts
            if contact.name
        }

    for row_number, row in enumerate(source_rows, start=2):
        try:
            mapped, custom_values = _map_row(row, CONTACT_COLUMN_MAP, definition_lookup)
            name = _clean_text(mapped.get("name"))
            if not name:
                summary.skipped += 1
                summary.errors.append(f"Satir {row_number}: Kisi adi zorunludur.")
                continue

            company = _resolve_company(mapped, company_by_name, company_by_id)
            if not company:
                summary.skipped += 1
                summary.errors.append(f"Satir {row_number}: Gecerli bir sirket bulunamadi.")
                continue

            payload = _build_contact_payload(mapped, custom_values, company.id)
            existing = contact_map.get((company.id, name.lower()))

            if existing:
                merged_custom_values = get_field_values("contact", existing.id)
                merged_custom_values.update(payload.pop("custom_values"))
                payload["custom_values"] = merged_custom_values
                update_contact(existing.id, payload)
                summary.updated += 1
            else:
                contact = create_contact(payload)
                contact_map[(company.id, name.lower())] = contact
                summary.added += 1
        except Exception as exc:
            summary.errors.append(f"Satir {row_number}: {exc}")

    return summary


def _read_tabular_file(file_path: str) -> list[dict[str, Any]]:
    target = Path(file_path)
    suffix = target.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(target)
    return _read_csv(target)


def _read_csv(target: Path) -> list[dict[str, Any]]:
    with target.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _read_xlsx(target: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(target, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue
        result.append({headers[index]: row[index] for index in range(len(headers))})
    return result


def _build_definition_lookup(definitions: list[FieldDefinition]) -> dict[str, FieldDefinition]:
    lookup: dict[str, FieldDefinition] = {}
    for definition in definitions:
        lookup[_normalize_header(definition.label)] = definition
        lookup[_normalize_header(definition.field_key)] = definition
    return lookup


def _map_row(
    row: dict[str, Any],
    known_columns: dict[str, str],
    definition_lookup: dict[str, FieldDefinition],
) -> tuple[dict[str, Any], dict[str, Any]]:
    mapped: dict[str, Any] = {}
    custom_values: dict[str, Any] = {}

    for header, value in row.items():
        normalized = _normalize_header(header)
        if not normalized:
            continue

        if normalized in known_columns:
            mapped[known_columns[normalized]] = value
            continue

        definition = definition_lookup.get(normalized)
        if definition:
            parsed_value = _convert_dynamic_value(value, definition.field_type)
            if parsed_value not in (None, ""):
                custom_values[definition.field_key] = parsed_value

    return mapped, custom_values


def _build_company_payload(mapped: dict[str, Any], custom_values: dict[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "name": _clean_text(mapped.get("name")) or "",
        "custom_values": custom_values,
    }

    if _clean_text(mapped.get("country")):
        payload["country"] = _clean_text(mapped.get("country"))
    if _clean_text(mapped.get("city")):
        payload["city"] = _clean_text(mapped.get("city"))
    if _clean_text(mapped.get("website")):
        payload["website"] = _clean_text(mapped.get("website"))
    if _clean_text(mapped.get("linkedin")):
        payload["linkedin"] = _clean_text(mapped.get("linkedin"))
    if mapped.get("priority") not in (None, ""):
        payload["priority"] = _parse_int(mapped.get("priority"), default=3)

    return payload


def _build_contact_payload(
    mapped: dict[str, Any],
    custom_values: dict[str, Any],
    company_id: int,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "name": _clean_text(mapped.get("name")) or "",
        "company_id": company_id,
        "custom_values": custom_values,
    }

    if _clean_text(mapped.get("title")):
        payload["title"] = _clean_text(mapped.get("title"))
    if _clean_text(mapped.get("email")):
        payload["email"] = _clean_text(mapped.get("email"))
    if _clean_text(mapped.get("phone")):
        payload["phone"] = _clean_text(mapped.get("phone"))
    if _clean_text(mapped.get("linkedin")):
        payload["linkedin"] = _clean_text(mapped.get("linkedin"))

    return payload


def _resolve_company(
    mapped: dict[str, Any],
    company_by_name: dict[str, Company],
    company_by_id: dict[int, Company],
) -> Company | None:
    company_id = mapped.get("company_id")
    if company_id not in (None, ""):
        parsed_id = _parse_int(company_id, default=0)
        if parsed_id in company_by_id:
            return company_by_id[parsed_id]

    company_name = _clean_text(mapped.get("company_name"))
    if not company_name:
        return None
    return company_by_name.get(company_name.lower())


def _convert_dynamic_value(value: Any, field_type: str) -> Any:
    if value in (None, ""):
        return None

    if field_type == "number":
        return _parse_float(value)
    if field_type == "boolean":
        return _parse_bool(value)
    if field_type == "date":
        return _parse_date(value)
    return _clean_text(value)


def _normalize_header(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return "".join(character for character in text.lower() if character.isalnum())


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(str(value).replace(",", ".")))
    except (TypeError, ValueError):
        return default


def _parse_float(value: Any) -> float | None:
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return _clean_text(value).lower() in {"1", "true", "evet", "yes", "x"}


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _clean_text(value)
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None
